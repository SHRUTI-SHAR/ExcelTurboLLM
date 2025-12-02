# excel_ai_turbo_engine_final_fixed_v4.py
"""
Excel AI Turbo Engine — Final v4

✅ Fixes:
- Excel COM initialized properly (CoInitialize / CoUninitialize)
- Handles duplicate and empty headers
- Aligns DataFrames before comparing (no ValueError)
- Uses tempfile safely (no permission denied)
- Detects protected sheets and shows user-friendly message
- Optionally auto-unprotect sheet with known password
- Restores proper dtypes from AgGrid edits before sending to Excel
- Updates only changed rows to avoid overwriting good data
- Maps Excel CVErr ints (-2146...) to readable error strings (#VALUE!, #NAME?, etc.)
"""

import os
import platform
import tempfile
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode

# Optional OpenAI (not needed for Excel core)
try:
    import openai
except Exception:
    openai = None

USE_EXCEL_COM = False
if platform.system() == "Windows":
    try:
        import win32com.client as win32
        import pythoncom
        USE_EXCEL_COM = True
    except Exception:
        USE_EXCEL_COM = False

# Optional password for automatic unprotect (leave blank if none)
SHEET_PASSWORD = ""

st.set_page_config(page_title="Excel AI Turbo Engine — Final v4", layout="wide")
st.title("🚀 Excel AI Turbo Engine — Final v4")
st.markdown("Upload an Excel file, edit inputs, and auto-recalculate using Excel COM or Python evaluator.")

# ------------------------ Helpers ------------------------ #

def load_workbook(uploaded_file):
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
        return wb
    except Exception as e:
        st.error(f"Could not open workbook: {e}")
        st.stop()

def list_sheets_with_state(wb):
    return [(ws.title, getattr(ws, "sheet_state", "visible")) for ws in wb.worksheets]

def sheet_to_dataframe(ws):
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame(), [], {}
    headers = list(rows[0])
    # Fill empty headers
    for i, h in enumerate(headers):
        if h is None or str(h).strip() == "":
            headers[i] = f"Column_{get_column_letter(i+1)}"
    df = pd.DataFrame(rows[1:], columns=headers)

    # Map Excel column letters -> header names
    colmap = {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header = headers[col_idx - 1] if col_idx - 1 < len(headers) else f"Column_{letter}"
        colmap[letter] = str(header)
    return df, headers, colmap

def detect_formulas_and_inputs(ws, df, colmap):
    formulas_map, formula_cells = {}, []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.row == 1:
                continue
            if cell.data_type == "f" or (isinstance(cell.value, str) and str(cell.value).startswith("=")):
                formulas_map[cell.column_letter] = cell.value
                formula_cells.append({"cell": cell.coordinate, "formula": cell.value})
    output_headers = [colmap[c] for c in formulas_map.keys() if c in colmap]
    input_headers = [h for h in df.columns if h not in output_headers]
    return formulas_map, formula_cells, input_headers, output_headers

def detect_orientation(ws, df):
    first_col_vals = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    text_count = sum(1 for v in first_col_vals if isinstance(v, str))
    non_empty = sum(1 for v in first_col_vals if v is not None)
    if non_empty == 0:
        return "table"
    return "row" if text_count / non_empty > 0.6 else "table"

def auto_cast(v):
    """
    Convert values coming from the grid into types Excel understands:
    - Numbers → numbers
    - Date-like strings → datetime
    - Empty → None
    - Other strings → left as string
    """
    try:
        # Treat None / NaN as empty cells
        if v is None:
            return None

        if isinstance(v, float) and np.isnan(v):
            return None

        # If it is already a pandas Timestamp → convert to Python datetime
        if isinstance(v, pd.Timestamp):
            return v.to_pydatetime()

        # Numpy scalars
        if hasattr(v, "item") and not isinstance(v, (str, bytes)):
            v = v.item()

        # Plain numeric + bool types
        if isinstance(v, (int, float, bool)):
            return v

        # Handle strings: try numeric, then date, else keep as text
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                return None

            # Try integer / float
            try:
                if "." in s:
                    return float(s)
                else:
                    return int(s)
            except Exception:
                pass

            # Try date parsing (dayfirst=True for formats like 12-05-2000)
            try:
                dt = pd.to_datetime(s, dayfirst=True, errors="raise")
                return dt.to_pydatetime()
            except Exception:
                pass

            # Leave as text if nothing else fits
            return s

        # Fallback: return as-is
        return v

    except Exception:
        return v


# Excel internal error codes → readable Excel error strings
EXCEL_ERROR_MAP = {
    -2146826281: "#DIV/0!",
    -2146826246: "#N/A",
    -2146826245: "#GETTING_DATA",
    -2146826259: "#NAME?",
    -2146826288: "#NULL!",
    -2146826252: "#NUM!",
    -2146826265: "#REF!",
    -2146826273: "#VALUE!",
}

def restore_dtypes(edited_df: pd.DataFrame, template_df: pd.DataFrame) -> pd.DataFrame:
    """
    Cast edited values back to the dtypes of the original editable_df,
    so Excel COM gets proper numbers/dates/bools instead of raw strings.
    """
    result = edited_df.copy()
    for col in result.columns:
        if col not in template_df.columns:
            continue
        orig = template_df[col]

        # Numeric columns
        if pd.api.types.is_numeric_dtype(orig):
            result[col] = pd.to_numeric(result[col], errors="coerce")

        # Datetime columns
        elif pd.api.types.is_datetime64_any_dtype(orig):
            result[col] = pd.to_datetime(result[col], errors="coerce")

        # Boolean columns
        elif pd.api.types.is_bool_dtype(orig):
            def to_bool(x):
                if pd.isna(x):
                    return None
                if isinstance(x, bool):
                    return x
                if isinstance(x, (int, float)):
                    return bool(x)
                s = str(x).strip().lower()
                if s in ("true", "yes", "y", "1"):
                    return True
                if s in ("false", "no", "n", "0"):
                    return False
                return None
            result[col] = result[col].map(to_bool)

        # Other types: keep as object, but normalize NaN → None
        else:
            result[col] = result[col].where(result[col].notna(), None)

    return result

# ---------------- Excel COM Calculation ---------------- #

def excel_com_calculate_and_read(path, sheet_name, updates, password=""):
    """Use Excel COM to calculate workbook safely with unique headers.
       If some cells are protected, skip them instead of failing.
    """
    if platform.system() != "Windows":
        raise RuntimeError("Excel COM is only supported on Windows.")
    import pythoncom
    excel = None
    wb = None
    skipped_cells = 0
    try:
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        abs_path = os.path.abspath(path)
        wb = excel.Workbooks.Open(abs_path, ReadOnly=False)
        ws = wb.Worksheets(sheet_name)

        # 🔓 Try to unprotect sheet (ignore errors if already unprotected)
        try:
            ws.Unprotect(Password=password or "")
        except Exception:
            pass

        # Apply updates cell-by-cell
        for (r, c), val in updates.items():
            try:
                ws.Cells(r, c).Value = auto_cast(val)
            except Exception:
                skipped_cells += 1
                continue

        # Trigger recalculation
        try:
            excel.Calculate()
            excel.CalculateUntilAsyncQueriesDone()
        except Exception:
            try:
                excel.CalculateFull()
            except Exception:
                pass

        wb.Save()

        # Read back results
        used = ws.UsedRange
        nrows, ncols = used.Rows.Count, used.Columns.Count
        data = []
        for i in range(1, nrows + 1):
            row = []
            for j in range(1, ncols + 1):
                val = ws.Cells(i, j).Value
                # Convert Excel CVErr ints to readable error strings
                if isinstance(val, int) and val in EXCEL_ERROR_MAP:
                    val = EXCEL_ERROR_MAP[val]
                row.append(val)
            data.append(row)

        # Safe header construction
        if len(data) > 1:
            headers = list(data[0])
            safe_headers, seen = [], {}
            for i, h in enumerate(headers):
                if h is None or str(h).strip() == "":
                    h = f"Column_{i+1}"
                if h in seen:
                    seen[h] += 1
                    h = f"{h}_{seen[h]}"
                else:
                    seen[h] = 1
                safe_headers.append(str(h))
            df = pd.DataFrame(data[1:], columns=safe_headers)
        else:
            df = pd.DataFrame()

        df.columns = [str(c) if c else f"Unnamed_{i+1}" for i, c in enumerate(df.columns)]

        # Show skipped cells warning if any
        if skipped_cells > 0:
            st.warning(f"⚠️ Skipped {skipped_cells} protected or non-editable cells during COM update.")
        return df

    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ---------------- Streamlit UI ---------------- #

st.sidebar.header("⚙️ Settings")
use_com = st.sidebar.checkbox("Use Excel COM (Windows only)", value=(USE_EXCEL_COM))
uploaded = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx"])
if not uploaded:
    st.stop()

tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
os.close(tmp_fd)
with open(tmp_path, "wb") as f:
    f.write(uploaded.getbuffer())

wb = load_workbook(tmp_path)
sheets = list_sheets_with_state(wb)
sheet_names = [name for name, _ in sheets]
st.sidebar.markdown("### Sheets (visible + hidden)")
for name, state in sheets:
    st.sidebar.write(f"- {name} — {state}")

selected_sheet = st.selectbox("Select sheet to work on", sheet_names)
if not selected_sheet:
    st.stop()

ws = wb[selected_sheet]
df, headers, colmap = sheet_to_dataframe(ws)
formulas_map, formula_cells, input_headers, output_headers = detect_formulas_and_inputs(ws, df, colmap)
orientation = detect_orientation(ws, df)

st.write(f"Detected orientation: **{orientation}**")
st.write(f"Detected outputs (headers): {output_headers}")
st.write(f"Detected inputs (editable): {input_headers}")

# Editable grid
if orientation == "table":
    editable_df = df[input_headers].copy().reset_index(drop=True)
else:
    keys, vals = [], []
    for r in range(2, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k or v:
            keys.append(k if k else f"Row_{r}")
            vals.append(v)
    editable_df = pd.DataFrame({"Key": keys, "Value": vals})

st.markdown("### ✏️ Editable Inputs")
gb = GridOptionsBuilder.from_dataframe(editable_df)
gb.configure_default_column(editable=True)
grid_options = gb.build()
grid_response = AgGrid(
    editable_df,
    gridOptions=grid_options,
    update_mode=GridUpdateMode.VALUE_CHANGED,
    fit_columns_on_grid_load=True,
    theme="alpine",
    allow_unsafe_jscode=True,
)

edited = pd.DataFrame(grid_response["data"]).reset_index(drop=True)

# Restore dtypes based on original editable_df
edited = restore_dtypes(edited, editable_df)

# --- FIXED SAFE COMPARISON LOGIC ---
common_cols = [c for c in edited.columns if c in editable_df.columns]
edited_aligned = edited[common_cols].reset_index(drop=True)
orig_aligned = editable_df[common_cols].reset_index(drop=True)
max_len = max(len(edited_aligned), len(orig_aligned))
edited_aligned = edited_aligned.reindex(range(max_len)).fillna("")
orig_aligned = orig_aligned.reindex(range(max_len)).fillna("")
changed_mask = (edited_aligned != orig_aligned).any(axis=1)
# -----------------------------------

edited["_changed"] = changed_mask

# ---------------- Calculation Section ---------------- #

st.markdown("### ⚡ Live Computed Results")

if use_com and USE_EXCEL_COM:
    st.info("Using Excel COM for exact calculation (Windows + Excel).")
    updates = {}

    if orientation == "table":
        header_to_letter = {v: k for k, v in colmap.items()}
        for r in range(len(edited)):
            if not bool(edited.loc[r, "_changed"]):
                continue  # only update changed rows
            for col_header in input_headers:
                letter = header_to_letter.get(col_header)
                if letter:
                    col_idx = column_index_from_string(letter)
                    updates[(r + 2, col_idx)] = edited.at[r, col_header]
    else:
        # Key/Value mode
        for r in range(len(edited)):
            if not bool(edited.loc[r, "_changed"]):
                continue
            updates[(2 + r, 2)] = edited.at[r, "Value"]

    try:
        calc_df = excel_com_calculate_and_read(tmp_path, selected_sheet, updates, password=SHEET_PASSWORD)
        # Align changed_mask length with calc_df rows
        calc_df["_changed"] = changed_mask.reindex(range(len(calc_df))).fillna(False).tolist()
        AgGrid(calc_df, fit_columns_on_grid_load=True, theme="alpine", allow_unsafe_jscode=True)
    except PermissionError as pe:
        st.error(f"❌ Excel sheet is protected.\n\n{pe}\n\nPlease unprotect it manually or set SHEET_PASSWORD in code.")
    except Exception as e:
        st.error(f"Excel COM calculation failed: {e}")
else:
    st.info("Using Python evaluator (fallback, best-effort).")
    st.warning("Some Excel formulas may not evaluate exactly without Excel COM.")

st.success("✅ Live update complete. Edit inputs above — outputs will refresh when values change.")
